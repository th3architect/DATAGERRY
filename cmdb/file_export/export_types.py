# DATAGERRY - OpenSource Enterprise CMDB
# Copyright (C) 2019 NETHINKS GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import csv
import io
import json
import re
import tempfile
import xml.dom.minidom
import xml.etree.ElementTree as ET
import zipfile
import openpyxl
from cmdb.utils.helpers import load_class

from cmdb.database.managers import DatabaseManagerMongo
from cmdb.framework.managers.type_manager import TypeManager
from cmdb.framework.cmdb_object_manager import CmdbObjectManager
from cmdb.utils import json_encoding
from cmdb.utils.system_config import SystemConfigReader

database_manager = DatabaseManagerMongo(**SystemConfigReader().get_all_values_from_section('Database'))
object_manager = CmdbObjectManager(database_manager=database_manager)
type_manager = TypeManager(database_manager=database_manager)


class ExportType:
    FILE_EXTENSION = None
    LABEL = None
    MULTITYPE_SUPPORT = False
    ICON = None
    DESCRIPTION = None
    ACTIVE = None

    def __init__(self):
        pass

    def export(self, object_list, *args):
        pass


class ZipExportType(ExportType):

    FILE_EXTENSION = "zip"
    LABEL = "ZIP"
    MULTITYPE_SUPPORT = True
    ICON = "file-archive"
    DESCRIPTION = "Export Zipped Files"
    ACTIVE = True

    def export(self, object_list, *args):

        """
        Export a zip file, containing the object list sorted by type in several files.

        Args:
            object_list: List of objects to be exported
            args: the filetype with which the objects are stored

        Returns:
            zip file containing object files separated by types
        """

        # check what export type is requested and intitializes a new zip file in memory
        export_type = load_class("cmdb.file_export.export_types." + args[0])()
        zipped_file = io.BytesIO()

        # Build .zip file
        with zipfile.ZipFile(zipped_file, "a", zipfile.ZIP_DEFLATED, False) as f:

            # Enters loop until the object list is empty
            while len(object_list) > 0:

                # Set what type the loop filters to and makes an empty list
                type_id = object_list[len(object_list) - 1].type_id
                type_name = object_manager.get_type(type_id).get_name()
                type_list = []

                # Filters object list to the current type_id and inserts it into type_list
                # When an object is inserted into type_list it gets deleted from object_list
                for i in range(len(object_list) - 1, -1, -1):
                    if object_list[i].type_id == type_id:
                        type_list.append(object_list[i])
                        del object_list[i]

                # Runs the requested export function and returns the output in the export variable
                export = export_type.export(type_list)

                # check if export output is a string, bytes or a file and inserts it into the zip file
                if isinstance(export, str) or isinstance(export, bytes):
                    f.writestr((type_name + "_ID_" + str(type_id) + "." + export_type.FILE_EXTENSION).format(i), export)
                else:
                    f.writestr((type_name + "_ID_" + str(type_id) + "." + export_type.FILE_EXTENSION).format(i), export.getvalue())

        # returns zipped file
        zipped_file.seek(0)
        return zipped_file


class CsvExportType(ExportType):

    FILE_EXTENSION = "csv"
    LABEL = "CSV"
    MULTITYPE_SUPPORT = False
    ICON = "file-csv"
    DESCRIPTION = "Export as CSV (only of the same type)"
    ACTIVE = True

    def export(self, object_list, *args):

        """ Exports object_list as .csv file

        Args:
            object_list: The objects to be exported
            args:

        Returns:
            Csv file containing the object_list
        """

        # init values
        header = ['public_id', 'active']
        rows = []
        current_type_id = None
        current_type_fields = []

        for obj in object_list:
            # get type from first object and setup csv header
            if current_type_id is None:
                current_type_id = obj.type_id
                current_type_fields = type_manager.get(obj.type_id).get_fields()
                for type_field in current_type_fields:
                    header.append(type_field.get('name'))

            # throw Exception if objects of different type are detected
            if current_type_id != obj.type_id:
                raise Exception({'message': 'CSV can export only object of the same type'})

            # get object fields as dict:
            obj_fields_dict = {}
            for obj_field in obj.fields:
                obj_field_name = obj_field.get('name')
                obj_fields_dict[obj_field_name] = obj_field.get('value')

            # define output row
            row = []
            row.append(str(obj.public_id))
            row.append(str(obj.active))
            for type_field in current_type_fields:
                row.append(str(obj_fields_dict.get(type_field.get('name'), None)))
            rows.append(row)

        return self.csv_writer(header, rows)

    def csv_writer(self, header, rows, dialect=csv.excel):

        csv_file = io.StringIO()
        writer = csv.writer(csv_file, dialect=dialect)
        writer.writerow(header)
        for row in rows:
            writer.writerow(row)
        csv_file.seek(0)
        return csv_file


class JsonExportType(ExportType):

    FILE_EXTENSION = "json"
    LABEL = "JSON"
    MULTITYPE_SUPPORT = True
    ICON = "file-code"
    DESCRIPTION = "Export as JSON"
    ACTIVE = True

    def export(self, object_list, *args):

        """Exports object_list as .json file

        Args:
            object_list: The objects to be exported
            args:

        Returns:
            Json file containing the object_list
        """

        output = []

        for obj in object_list:
            # init output element
            output_element = {}
            output_element['public_id'] = obj.public_id
            output_element['active'] = obj.active
            output_element['type'] = type_manager.get(obj.type_id).get_label()
            output_element['fields'] = []

            # get object fields as dict:
            obj_fields_dict = {}
            for obj_field in obj.fields:
                obj_field_name = obj_field.get('name')
                obj_fields_dict[obj_field_name] = obj_field.get('value')

            # walk over all type fields and add object field values
            for type_field in type_manager.get(obj.type_id).get_fields():
                output_element['fields'].append({
                    'name': type_field.get('name'),
                    'value': obj_fields_dict.get(type_field.get('name'), None)
                })

            output.append(output_element)

        return json.dumps(output, default=json_encoding.default, ensure_ascii=False, indent=2)


class XlsxExportType(ExportType):

    FILE_EXTENSION = "xlsx"
    LABEL = "XLSX"
    MULTITYPE_SUPPORT = True
    ICON = "file-excel"
    DESCRIPTION = "Export as XLS"
    ACTIVE = True

    def export(self, object_list, *args):

        """Exports object_list as .xlsx file

        Args:
            object_list: The objects to be exported
            args:

        Returns:
            Xlsx file containing the object_list
        """

        workbook = self.create_xls_object(object_list)

        # save workbook
        with tempfile.NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            return tmp.read()

    def create_xls_object(self, object_list):

        # create workbook
        workbook = openpyxl.Workbook()

        # sorts data_list by type_id
        type_id = "type_id"
        decorated = [(dict_.__dict__[type_id], dict_.__dict__) for dict_ in object_list]
        decorated = sorted(decorated, key=lambda x: x[0], reverse=False)
        sorted_list = [dict_ for (key, dict_) in decorated]

        # init values
        current_type_id = None
        current_type_fields = []
        i = 0
        for obj in sorted_list:

            # check, if starting a new object type
            if current_type_id != obj[type_id]:
                # set current type id and fields
                current_type_id = obj[type_id]
                current_type_fields = type_manager.get(obj[type_id]).get_fields()
                i = 1

                # delete default sheet
                workbook.remove(workbook.active)

                # start a new worksheet and rename it
                title = self.__normalize_sheet_title(object_manager.get_type(obj[type_id]).label)
                sheet = workbook.create_sheet(title)

                # insert header: public_id, active
                cell = sheet.cell(row=i, column=1)
                cell.value = 'public_id'
                cell = sheet.cell(row=i, column=2)
                cell.value = 'active'

                # insert header: get fields from type definition
                c = 3
                for type_field in current_type_fields:
                    cell = sheet.cell(row=i, column=c)
                    cell.value = type_field.get('name')
                    c = c + 1
                i = i + 1

            # insert row values: public_id, active
            cell = sheet.cell(row=i, column=1)
            cell.value = str(obj["public_id"])
            cell = sheet.cell(row=i, column=2)
            cell.value = str(obj["active"])

            # get object fields as dict:
            obj_fields_dict = {}
            for obj_field in obj['fields']:
                obj_field_name = obj_field.get('name')
                obj_fields_dict[obj_field_name] = obj_field.get('value')

            # insert row values: fields
            c = 3
            for type_field in current_type_fields:
                cell = sheet.cell(row=i, column=c)
                cell.value = str(obj_fields_dict.get(type_field.get('name'), None))
                c = c + 1

            # increase row number
            i = i + 1

        return workbook

    def __normalize_sheet_title(self, input_data):
        return re.sub('[\\*?:/\[\]]', '_', input_data)


class XmlExportType(ExportType):

    FILE_EXTENSION = "xml"
    LABEL = "XML"
    MULTITYPE_SUPPORT = True
    ICON = "file-alt"
    DESCRIPTION = "Export as XML"
    ACTIVE = True

    def export(self, object_list, *args):

        """Exports object_list as .xml file

        Args:
            object_list: The objects to be exported
            args:

        Returns:
            Xml file containing the object_list
        """

        # object list
        cmdb_object_list = ET.Element('objects')

        for obj in object_list:
            # get object fields as dict:
            obj_fields_dict = {}
            for obj_field in obj.fields:
                obj_field_name = obj_field.get('name')
                obj_fields_dict[obj_field_name] = obj_field.get('value')

            # xml output: object
            cmdb_object = ET.SubElement(cmdb_object_list, 'object')
            cmdb_object_meta = ET.SubElement(cmdb_object, 'meta')
            # xml output meta: public
            cmdb_object_meta_id = ET.SubElement(cmdb_object_meta, 'public_id')
            cmdb_object_meta_id.text = str(obj.public_id)
            # xml output meta: active
            cmdb_object_meta_active = ET.SubElement(cmdb_object_meta, 'active')
            cmdb_object_meta_active.text = str(obj.active)
            # xml output meta: type
            cmdb_object_meta_type = ET.SubElement(cmdb_object_meta, 'type')
            cmdb_object_meta_type.text = object_manager.get_type(obj.type_id).label
            # xml output: fields
            cmdb_object_fields = ET.SubElement(cmdb_object, 'fields')
            # walk over all type fields and add object field values
            for type_field in type_manager.get(obj.type_id).get_fields():
                field_attribs = {}
                field_attribs["name"] = str(type_field.get('name'))
                field_attribs["value"] = str(obj_fields_dict.get(type_field.get('name'), None))
                ET.SubElement(cmdb_object_fields, "field", field_attribs)

        # return xml as string (pretty printed)
        return xml.dom.minidom.parseString(
            ET.tostring(cmdb_object_list, encoding='unicode', method='xml')).toprettyxml()
